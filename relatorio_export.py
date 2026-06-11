import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import database as db


def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value, bg="2C3E50"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _thin()
    return c


def _auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 60)


STATUS_CORES = {
    "novo":        "FFF9C4",
    "em_cotacao":  "FFE0B2",
    "aprovado":    "B3E5FC",
    "em_execucao": "BBDEFB",
    "concluido":   "C8E6C9",
    "cancelado":   "FFCDD2",
}


def gerar_excel(cond_id=None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Demandas sheet ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Demandas")
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Relatório de Demandas — {date.today()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    hdrs = ["#", "Condomínio", "Título", "Categoria", "Prestador",
            "Valor R$", "Prazo", "Status"]
    for c, h in enumerate(hdrs, 1):
        _header_cell(ws, 2, c, h)

    demandas = db.listar_demandas(condominio_id=cond_id)
    for r, d in enumerate(demandas, 3):
        cond_nome = (d.get("condominios") or {}).get("nome", "")
        vals = [
            d["id"], cond_nome, d["titulo"],
            d["categoria"].capitalize(),
            d.get("prestador") or "—",
            d.get("valor") or "",
            d.get("data_limite") or "—",
            d["status"].replace("_", " ").capitalize(),
        ]
        cor = STATUS_CORES.get(d["status"], "FFFFFF")
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.border = _thin()
            if c == 6 and v:
                cell.number_format = 'R$ #,##0.00'
    _auto_width(ws)

    # ── Manutenções sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Manutenções")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"Relatório de Manutenções — {date.today()}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center")

    hdrs2 = ["#", "Condomínio", "Tipo", "Prestador",
             "Data Realização", "Vencimento", "Valor R$", "Status"]
    for c, h in enumerate(hdrs2, 1):
        _header_cell(ws2, 2, c, h)

    manuts = db.listar_manutencoes(condominio_id=cond_id)
    hoje = date.today()
    for r, m in enumerate(manuts, 3):
        cond_nome = (m.get("condominios") or {}).get("nome", "")
        if m.get("data_vencimento"):
            diff = (date.fromisoformat(m["data_vencimento"]) - hoje).days
            if diff < 0:
                status_m, cor = f"Vencido {abs(diff)}d", "FFCDD2"
            elif diff <= db.DIAS_ALERTA_MANUTENCAO:
                status_m, cor = f"Vence em {diff}d", "FFF9C4"
            else:
                status_m, cor = "OK", "C8E6C9"
        else:
            status_m, cor = "—", "FFFFFF"

        vals2 = [
            m["id"], cond_nome, m["tipo"],
            m.get("prestador") or "—",
            m.get("data_realizacao") or "—",
            m.get("data_vencimento") or "—",
            m.get("valor") or "",
            status_m,
        ]
        for c, v in enumerate(vals2, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.border = _thin()
            if c == 7 and v:
                cell.number_format = 'R$ #,##0.00'
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_pdf(cond_id=None) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Relatório — Síndico App", align="C")
    pdf.ln(5)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Gerado em: {date.today()}", align="C")
    pdf.ln(10)

    # ── Demandas ───────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Demandas", fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(1)

    cols  = ["#", "Condomínio", "Título", "Categoria", "Status"]
    widths= [10, 45, 65, 30, 35]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(240, 242, 248)
    for h, w in zip(cols, widths):
        pdf.cell(w, 7, h, border=1, fill=True)
    pdf.ln()

    demandas = db.listar_demandas(condominio_id=cond_id)
    pdf.set_font("Helvetica", "", 8)
    STATUS_RGB = {
        "novo": (255, 249, 196), "em_cotacao": (255, 224, 178),
        "aprovado": (179, 229, 252), "em_execucao": (187, 222, 251),
        "concluido": (200, 230, 201), "cancelado": (255, 205, 210),
    }
    for d in demandas:
        rgb = STATUS_RGB.get(d["status"], (255, 255, 255))
        pdf.set_fill_color(*rgb)
        cond_nome = (d.get("condominios") or {}).get("nome", "")
        vals = [
            str(d["id"]),
            cond_nome[:22],
            d["titulo"][:35],
            d["categoria"].capitalize(),
            d["status"].replace("_", " ").capitalize(),
        ]
        for v, w in zip(vals, widths):
            pdf.cell(w, 6, v, border=1, fill=True)
        pdf.ln()

    pdf.ln(6)

    # ── Manutenções ────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Manutenções", fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(1)

    cols2  = ["#", "Condomínio", "Tipo", "Prestador", "Vencimento", "Status"]
    widths2= [10, 42, 38, 38, 28, 29]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(240, 242, 248)
    for h, w in zip(cols2, widths2):
        pdf.cell(w, 7, h, border=1, fill=True)
    pdf.ln()

    manuts = db.listar_manutencoes(condominio_id=cond_id)
    hoje = date.today()
    pdf.set_font("Helvetica", "", 8)
    for m in manuts:
        cond_nome = (m.get("condominios") or {}).get("nome", "")
        if m.get("data_vencimento"):
            diff = (date.fromisoformat(m["data_vencimento"]) - hoje).days
            if diff < 0:
                status_m, rgb = f"Vencido {abs(diff)}d", (255, 205, 210)
            elif diff <= db.DIAS_ALERTA_MANUTENCAO:
                status_m, rgb = f"Em {diff}d", (255, 249, 196)
            else:
                status_m, rgb = "OK", (200, 230, 201)
        else:
            status_m, rgb = "—", (255, 255, 255)
        pdf.set_fill_color(*rgb)
        vals2 = [
            str(m["id"]),
            cond_nome[:22],
            m["tipo"][:20],
            (m.get("prestador") or "—")[:20],
            m.get("data_vencimento") or "—",
            status_m,
        ]
        for v, w in zip(vals2, widths2):
            pdf.cell(w, 6, v, border=1, fill=True)
        pdf.ln()

    return bytes(pdf.output())
